from datetime import time

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font


def export_to_excel(queryset):
    """Creates an Excel file from the given queryset and returns an HTTP response."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=registration_details.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    columns = ['registration_date',
               'interest',
               'user__first_name',
               'user__last_name',
               'user__email',
               'user__userdetails__mobile',
               'user__userdetails__gender',
               'user__userdetails__dob',
               'user__userdetails__designation',
               'user__userdetails__organization']

    columns_to_excel = ['Registration Date', 'Participat Interest', 'First Name', 'Last Name', 'Email', 'Mobile Number',
                        'Gender', 'DOB', 'Designation', 'Organization']

    # Header row (optional styling for bold)
    for col_num, column in enumerate(columns_to_excel, 1):
        cell = worksheet.cell(row=1, column=col_num, value=column)
        cell.font = Font(bold=True)

    # Data rows
    for row_num, obj in enumerate(queryset, 2):
        for col_num, column in enumerate(columns, 1):
            value = obj[column]
            if isinstance(value, (timezone.datetime, time)):
                value = value.replace(tzinfo=None)
            worksheet.cell(row=row_num, column=col_num).value = value

    workbook.save(response)
    return response


def export_emails_for_newsletters(queryset):
    """Creates an Excel file from the given queryset and returns an HTTP response."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=newsletter_registration_details.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    columns = ['user__first_name', 'user__last_name', 'user__email']

    columns_to_excel = ['First Name', 'Last Name', 'Email']

    # Header row (optional styling for bold)
    for col_num, column in enumerate(columns_to_excel, 1):
        cell = worksheet.cell(row=1, column=col_num, value=column)
        cell.font = Font(bold=True)

    # Data rows
    for row_num, obj in enumerate(queryset, 2):
        for col_num, column in enumerate(columns, 1):
            value = obj[column]
            if isinstance(value, (timezone.datetime, time)):
                value = value.replace(tzinfo=None)
            worksheet.cell(row=row_num, column=col_num).value = value

    workbook.save(response)
    return response

def str_to_bool(value):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ['true', '1', 'yes', 'on']

def str_to_int(value, default=3):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default